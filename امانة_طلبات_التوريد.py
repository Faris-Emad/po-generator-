import webbrowser
import threading
import time
from flask import Flask, request, send_file, render_template_string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import os
import sys

app = Flask(__name__)

HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>نظام إنشاء طلبات التوريد - شركة الأمانة</title>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Cairo:wght@300;400;600;700;800&display=swap');
        
        * { margin: 0; padding: 0; box-sizing: border-box; }
        
        body {
            font-family: 'Cairo', sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
            direction: rtl;
        }
        
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
            animation: slideIn 0.5s ease-out;
        }
        
        @keyframes slideIn {
            from { opacity: 0; transform: translateY(-20px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        .header {
            background: linear-gradient(135deg, #4472C4 0%, #2a4d8f 100%);
            color: white;
            padding: 40px;
            text-align: center;
            position: relative;
            overflow: hidden;
        }
        
        .header::before {
            content: '';
            position: absolute;
            top: -50%;
            right: -50%;
            width: 200%;
            height: 200%;
            background: radial-gradient(circle, rgba(255,255,255,0.1) 0%, transparent 70%);
            animation: pulse 4s infinite;
        }
        
        @keyframes pulse {
            0%, 100% { transform: scale(1); }
            50% { transform: scale(1.1); }
        }
        
        .header h1 {
            font-size: 2.5em;
            font-weight: 800;
            margin-bottom: 10px;
            position: relative;
            z-index: 1;
        }
        
        .header p {
            font-size: 1.2em;
            font-weight: 300;
            position: relative;
            z-index: 1;
        }
        
        .form-container { padding: 40px; }
        
        .section {
            margin-bottom: 40px;
            animation: fadeIn 0.6s ease-out;
        }
        
        @keyframes fadeIn {
            from { opacity: 0; }
            to { opacity: 1; }
        }
        
        .section-title {
            font-size: 1.5em;
            font-weight: 700;
            color: #4472C4;
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 3px solid #4472C4;
            display: flex;
            align-items: center;
            gap: 10px;
        }
        
        .section-title::before {
            content: '●';
            font-size: 0.7em;
        }
        
        .form-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 20px;
        }
        
        .form-group { display: flex; flex-direction: column; }
        
        label {
            font-weight: 600;
            color: #333;
            margin-bottom: 8px;
            font-size: 0.95em;
        }
        
        input, select, textarea {
            padding: 12px 15px;
            border: 2px solid #e0e0e0;
            border-radius: 10px;
            font-family: 'Cairo', sans-serif;
            font-size: 1em;
            transition: all 0.3s ease;
            background: #f8f9fa;
        }
        
        input:focus, select:focus, textarea:focus {
            outline: none;
            border-color: #4472C4;
            background: white;
            box-shadow: 0 0 0 3px rgba(68, 114, 196, 0.1);
        }
        
        .items-section {
            background: #f8f9fa;
            padding: 30px;
            border-radius: 15px;
            margin-bottom: 30px;
        }
        
        .item-row {
            display: grid;
            grid-template-columns: 50px 150px 1fr 2fr 100px 120px auto;
            gap: 15px;
            align-items: end;
            margin-bottom: 15px;
            padding: 15px;
            background: white;
            border-radius: 10px;
            box-shadow: 0 2px 5px rgba(0,0,0,0.05);
            transition: transform 0.2s ease;
        }
        
        .item-row:hover { transform: translateX(-5px); }
        
        .add-item-btn, .remove-item-btn {
            padding: 12px 25px;
            border: none;
            border-radius: 10px;
            font-family: 'Cairo', sans-serif;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s ease;
            font-size: 1em;
        }
        
        .add-item-btn {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
        }
        
        .add-item-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(102, 126, 234, 0.6);
        }
        
        .remove-item-btn {
            background: #dc3545;
            color: white;
            padding: 8px 12px;
            font-size: 0.9em;
        }
        
        .remove-item-btn:hover {
            background: #c82333;
            transform: scale(1.05);
        }
        
        .totals {
            background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
            padding: 25px;
            border-radius: 15px;
            margin-top: 30px;
        }
        
        .total-row {
            display: flex;
            justify-content: space-between;
            padding: 12px 0;
            font-size: 1.1em;
            border-bottom: 1px solid #dee2e6;
        }
        
        .total-row:last-child {
            border-bottom: none;
            font-weight: 800;
            font-size: 1.3em;
            color: #4472C4;
            margin-top: 10px;
            padding-top: 15px;
            border-top: 3px solid #4472C4;
        }
        
        .submit-btn {
            width: 100%;
            padding: 20px;
            background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
            color: white;
            border: none;
            border-radius: 15px;
            font-size: 1.3em;
            font-weight: 700;
            cursor: pointer;
            font-family: 'Cairo', sans-serif;
            margin-top: 30px;
            box-shadow: 0 10px 30px rgba(40, 167, 69, 0.3);
            transition: all 0.3s ease;
        }
        
        .submit-btn:hover {
            transform: translateY(-3px);
            box-shadow: 0 15px 40px rgba(40, 167, 69, 0.5);
        }
        
        @media (max-width: 768px) {
            .item-row { grid-template-columns: 1fr; }
            .header h1 { font-size: 1.8em; }
            .form-grid { grid-template-columns: 1fr; }
        }
        
        .loading {
            display: none;
            text-align: center;
            padding: 20px;
        }
        
        .loading.active { display: block; }
        
        .spinner {
            border: 4px solid #f3f3f3;
            border-top: 4px solid #4472C4;
            border-radius: 50%;
            width: 50px;
            height: 50px;
            animation: spin 1s linear infinite;
            margin: 0 auto;
        }
        
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🏢 نظام إنشاء طلبات التوريد</h1>
            <p>شركة الأمانة للتوريدات العامة</p>
        </div>
        
        <div class="form-container">
            <form id="purchaseOrderForm">
                <div class="section">
                    <div class="section-title">بيانات الشركة</div>
                    <div class="form-grid">
                        <div class="form-group">
                            <label>رقم الطلب</label>
                            <input type="text" id="poNumber" value="PO-001" required>
                        </div>
                        <div class="form-group">
                            <label>التاريخ</label>
                            <input type="date" id="poDate" required>
                        </div>
                        <div class="form-group">
                            <label>الرقم الضريبي للشركة</label>
                            <input type="text" id="companyTaxId">
                        </div>
                        <div class="form-group">
                            <label>السجل التجاري</label>
                            <input type="text" id="commercialReg">
                        </div>
                    </div>
                </div>
                
                <div class="section">
                    <div class="section-title">بيانات المورد</div>
                    <div class="form-grid">
                        <div class="form-group">
                            <label>اسم المورد *</label>
                            <input type="text" id="supplierName" required>
                        </div>
                        <div class="form-group">
                            <label>الرقم الضريبي للمورد *</label>
                            <input type="text" id="supplierTaxId" required>
                        </div>
                        <div class="form-group">
                            <label>رقم التليفون</label>
                            <input type="tel" id="supplierPhone">
                        </div>
                        <div class="form-group">
                            <label>البريد الإلكتروني</label>
                            <input type="email" id="supplierEmail">
                        </div>
                        <div class="form-group" style="grid-column: 1 / -1;">
                            <label>العنوان</label>
                            <input type="text" id="supplierAddress">
                        </div>
                    </div>
                </div>
                
                <div class="section">
                    <div class="section-title">شروط التوريد</div>
                    <div class="form-grid">
                        <div class="form-group">
                            <label>مدة التوريد (بالأيام)</label>
                            <select id="deliveryPeriod">
                                <option value="7">7 أيام</option>
                                <option value="14">14 يوم</option>
                                <option value="21">21 يوم</option>
                                <option value="30" selected>30 يوم</option>
                                <option value="45">45 يوم</option>
                                <option value="60">60 يوم</option>
                            </select>
                        </div>
                        <div class="form-group">
                            <label>مكان التسليم</label>
                            <input type="text" id="deliveryLocation">
                        </div>
                        <div class="form-group">
                            <label>شروط الدفع</label>
                            <select id="paymentTerms">
                                <option value="نقداً">نقداً</option>
                                <option value="آجل 30 يوم" selected>آجل 30 يوم</option>
                                <option value="آجل 60 يوم">آجل 60 يوم</option>
                                <option value="آجل 90 يوم">آجل 90 يوم</option>
                                <option value="على دفعتين">على دفعتين</option>
                                <option value="على ثلاث دفعات">على ثلاث دفعات</option>
                            </select>
                        </div>
                    </div>
                </div>
                
                <div class="section">
                    <div class="section-title">أصناف الطلب</div>
                    <div class="items-section" id="itemsContainer"></div>
                    <button type="button" class="add-item-btn" onclick="addItemRow()">+ إضافة صنف جديد</button>
                </div>
                
                <div class="totals">
                    <div class="total-row">
                        <span>الإجمالي قبل الضريبة:</span>
                        <span id="subtotal">0.00 ج.م</span>
                    </div>
                    <div class="total-row">
                        <span>ضريبة القيمة المضافة (14%):</span>
                        <span id="taxAmount">0.00 ج.م</span>
                    </div>
                    <div class="total-row">
                        <span>الإجمالي النهائي:</span>
                        <span id="totalAmount">0.00 ج.م</span>
                    </div>
                </div>
                
                <div class="section">
                    <div class="form-group">
                        <label>ملاحظات</label>
                        <textarea id="notes" rows="4"></textarea>
                    </div>
                </div>
                
                <button type="submit" class="submit-btn">📄 إنشاء وتحميل طلب التوريد</button>
                
                <div class="loading" id="loading">
                    <div class="spinner"></div>
                    <p>جاري إنشاء الملف...</p>
                </div>
            </form>
        </div>
    </div>
    
    <script>
        document.getElementById('poDate').valueAsDate = new Date();
        let itemCount = 0;
        
        function addItemRow() {
            itemCount++;
            const container = document.getElementById('itemsContainer');
            const row = document.createElement('div');
            row.className = 'item-row';
            row.innerHTML = `
                <div style="text-align: center; font-weight: 600; padding-top: 12px;">${itemCount}</div>
                <input type="text" placeholder="كود الصنف" class="item-code">
                <input type="text" placeholder="اسم الصنف" class="item-name" required>
                <input type="text" placeholder="الوصف" class="item-description">
                <input type="number" placeholder="الكمية" class="item-quantity" min="1" required onchange="calculateTotals()">
                <input type="number" placeholder="السعر" class="item-price" min="0" step="0.01" required onchange="calculateTotals()">
                <button type="button" class="remove-item-btn" onclick="removeItemRow(this)">حذف</button>
            `;
            container.appendChild(row);
        }
        
        function removeItemRow(btn) {
            btn.closest('.item-row').remove();
            updateItemNumbers();
            calculateTotals();
        }
        
        function updateItemNumbers() {
            const rows = document.querySelectorAll('.item-row');
            itemCount = rows.length;
            rows.forEach((row, index) => {
                row.querySelector('div').textContent = index + 1;
            });
        }
        
        function calculateTotals() {
            let subtotal = 0;
            document.querySelectorAll('.item-row').forEach(row => {
                const quantity = parseFloat(row.querySelector('.item-quantity').value) || 0;
                const price = parseFloat(row.querySelector('.item-price').value) || 0;
                subtotal += quantity * price;
            });
            
            const tax = subtotal * 0.14;
            const total = subtotal + tax;
            
            document.getElementById('subtotal').textContent = subtotal.toFixed(2) + ' ج.م';
            document.getElementById('taxAmount').textContent = tax.toFixed(2) + ' ج.م';
            document.getElementById('totalAmount').textContent = total.toFixed(2) + ' ج.م';
        }
        
        addItemRow();
        
        document.getElementById('purchaseOrderForm').addEventListener('submit', async function(e) {
            e.preventDefault();
            document.getElementById('loading').classList.add('active');
            
            const formData = {
                poNumber: document.getElementById('poNumber').value,
                poDate: document.getElementById('poDate').value,
                companyTaxId: document.getElementById('companyTaxId').value,
                commercialReg: document.getElementById('commercialReg').value,
                supplierName: document.getElementById('supplierName').value,
                supplierTaxId: document.getElementById('supplierTaxId').value,
                supplierPhone: document.getElementById('supplierPhone').value,
                supplierEmail: document.getElementById('supplierEmail').value,
                supplierAddress: document.getElementById('supplierAddress').value,
                deliveryPeriod: document.getElementById('deliveryPeriod').value,
                deliveryLocation: document.getElementById('deliveryLocation').value,
                paymentTerms: document.getElementById('paymentTerms').value,
                notes: document.getElementById('notes').value,
                items: []
            };
            
            document.querySelectorAll('.item-row').forEach(row => {
                formData.items.push({
                    code: row.querySelector('.item-code').value,
                    name: row.querySelector('.item-name').value,
                    description: row.querySelector('.item-description').value,
                    quantity: parseFloat(row.querySelector('.item-quantity').value) || 0,
                    price: parseFloat(row.querySelector('.item-price').value) || 0
                });
            });
            
            try {
                const response = await fetch('/generate', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(formData)
                });
                
                const blob = await response.blob();
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = `طلب_توريد_${formData.poNumber}_${new Date().toISOString().split('T')[0]}.xlsx`;
                document.body.appendChild(a);
                a.click();
                document.body.removeChild(a);
                window.URL.revokeObjectURL(url);
                
                document.getElementById('loading').classList.remove('active');
                alert('✅ تم إنشاء الملف بنجاح!');
            } catch (error) {
                document.getElementById('loading').classList.remove('active');
                alert('❌ حدث خطأ في إنشاء الملف');
            }
        });
    </script>
</body>
</html>
'''

def create_formatted_excel(data):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "طلب توريد"
    sheet.sheet_view.rightToLeft = True
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # رأس الشركة
    sheet.merge_cells('A1:G1')
    sheet['A1'] = 'شركة الأمانة للتوريدات العامة'
    sheet['A1'].font = Font(name='Arial', size=18, bold=True, color='FFFFFF')
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
    sheet['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    sheet.merge_cells('A2:G2')
    sheet['A2'] = 'طلب توريد'
    sheet['A2'].font = Font(name='Arial', size=14, bold=True)
    sheet['A2'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
    sheet.row_dimensions[2].height = 25
    
    # بيانات الشركة
    sheet['A4'] = 'بيانات الشركة:'
    sheet['A4'].font = Font(name='Arial', size=11, bold=True, underline='single')
    sheet['A4'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet.merge_cells('A4:B4')
    
    sheet['A5'] = 'رقم الطلب:'
    sheet['A5'].font = Font(name='Arial', size=10, bold=True)
    sheet['A5'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet['B5'] = data['poNumber']
    sheet['B5'].font = Font(name='Arial', size=10, color='0000FF')
    sheet['B5'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet['B5'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    sheet['D5'] = 'التاريخ:'
    sheet['D5'].font = Font(name='Arial', size=10, bold=True)
    sheet['D5'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet['E5'] = data['poDate']
    sheet['E5'].font = Font(name='Arial', size=10, color='0000FF')
    sheet['E5'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet['E5'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    sheet['A6'] = 'الرقم الضريبي:'
    sheet['A6'].font = Font(name='Arial', size=10, bold=True)
    sheet['A6'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet['B6'] = data['companyTaxId']
    sheet['B6'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    sheet['B6'].alignment = Alignment(horizontal='right', readingOrder=2)
    
    sheet['D6'] = 'السجل التجاري:'
    sheet['D6'].font = Font(name='Arial', size=10, bold=True)
    sheet['D6'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet['E6'] = data['commercialReg']
    sheet['E6'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    sheet['E6'].alignment = Alignment(horizontal='right', readingOrder=2)
    
    # بيانات المورد
    sheet['A8'] = 'بيانات المورد:'
    sheet['A8'].font = Font(name='Arial', size=11, bold=True, underline='single')
    sheet['A8'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet.merge_cells('A8:B8')
    
    sheet['A9'] = 'اسم المورد:'
    sheet['A9'].font = Font(name='Arial', size=10, bold=True)
    sheet['A9'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet['B9'] = data['supplierName']
    sheet['B9'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    sheet['B9'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet.merge_cells('B9:C9')
    
    sheet['D9'] = 'الرقم الضريبي:'
    sheet['D9'].font = Font(name='Arial', size=10, bold=True)
    sheet['D9'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet['E9'] = data['supplierTaxId']
    sheet['E9'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    sheet['E9'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet.merge_cells('E9:F9')
    
    sheet['A10'] = 'العنوان:'
    sheet['A10'].font = Font(name='Arial', size=10, bold=True)
    sheet['A10'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet['B10'] = data['supplierAddress']
    sheet['B10'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    sheet['B10'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet.merge_cells('B10:F10')
    
    sheet['A11'] = 'رقم التليفون:'
    sheet['A11'].font = Font(name='Arial', size=10, bold=True)
    sheet['A11'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet['B11'] = data['supplierPhone']
    sheet['B11'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    sheet['B11'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet.merge_cells('B11:C11')
    
    sheet['D11'] = 'البريد الإلكتروني:'
    sheet['D11'].font = Font(name='Arial', size=10, bold=True)
    sheet['D11'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet['E11'] = data['supplierEmail']
    sheet['E11'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    sheet['E11'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet.merge_cells('E11:F11')
    
    # شروط التوريد
    sheet['A13'] = 'شروط التوريد:'
    sheet['A13'].font = Font(name='Arial', size=11, bold=True, underline='single')
    sheet['A13'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet.merge_cells('A13:B13')
    
    sheet['A14'] = 'مدة التوريد:'
    sheet['A14'].font = Font(name='Arial', size=10, bold=True)
    sheet['A14'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet['B14'] = data['deliveryPeriod']
    sheet['B14'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    sheet['B14'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet['C14'] = 'يوم من تاريخ الطلب'
    sheet['C14'].font = Font(name='Arial', size=9, italic=True)
    sheet['C14'].alignment = Alignment(horizontal='right', readingOrder=2)
    
    sheet['D14'] = 'مكان التسليم:'
    sheet['D14'].font = Font(name='Arial', size=10, bold=True)
    sheet['D14'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet['E14'] = data['deliveryLocation']
    sheet['E14'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    sheet['E14'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet.merge_cells('E14:F14')
    
    sheet['A15'] = 'شروط الدفع:'
    sheet['A15'].font = Font(name='Arial', size=10, bold=True)
    sheet['A15'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet['B15'] = data['paymentTerms']
    sheet['B15'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    sheet['B15'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet.merge_cells('B15:C15')
    
    # جدول الأصناف
    headers = ['م', 'كود الصنف', 'اسم الصنف', 'الوصف', 'الكمية', 'سعر الوحدة', 'الإجمالي']
    header_row = 17
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
        cell.border = thin_border
    
    # إضافة الأصناف
    for idx, item in enumerate(data['items'], start=18):
        sheet.cell(row=idx, column=1).value = idx - 17
        sheet.cell(row=idx, column=1).alignment = Alignment(horizontal='center', readingOrder=2)
        sheet.cell(row=idx, column=1).border = thin_border
        
        sheet.cell(row=idx, column=2).value = item['code']
        sheet.cell(row=idx, column=2).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        sheet.cell(row=idx, column=2).alignment = Alignment(horizontal='right', readingOrder=2)
        sheet.cell(row=idx, column=2).border = thin_border
        
        sheet.cell(row=idx, column=3).value = item['name']
        sheet.cell(row=idx, column=3).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        sheet.cell(row=idx, column=3).alignment = Alignment(horizontal='right', readingOrder=2)
        sheet.cell(row=idx, column=3).border = thin_border
        
        sheet.cell(row=idx, column=4).value = item['description']
        sheet.cell(row=idx, column=4).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        sheet.cell(row=idx, column=4).alignment = Alignment(horizontal='right', readingOrder=2)
        sheet.cell(row=idx, column=4).border = thin_border
        
        sheet.cell(row=idx, column=5).value = item['quantity']
        sheet.cell(row=idx, column=5).font = Font(name='Arial', size=10, color='0000FF')
        sheet.cell(row=idx, column=5).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        sheet.cell(row=idx, column=5).alignment = Alignment(horizontal='center', readingOrder=2)
        sheet.cell(row=idx, column=5).border = thin_border
        
        sheet.cell(row=idx, column=6).value = item['price']
        sheet.cell(row=idx, column=6).font = Font(name='Arial', size=10, color='0000FF')
        sheet.cell(row=idx, column=6).number_format = '#,##0.00'
        sheet.cell(row=idx, column=6).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        sheet.cell(row=idx, column=6).alignment = Alignment(horizontal='right', readingOrder=2)
        sheet.cell(row=idx, column=6).border = thin_border
        
        sheet.cell(row=idx, column=7).value = f'=E{idx}*F{idx}'
        sheet.cell(row=idx, column=7).number_format = '#,##0.00'
        sheet.cell(row=idx, column=7).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        sheet.cell(row=idx, column=7).alignment = Alignment(horizontal='right', readingOrder=2)
        sheet.cell(row=idx, column=7).border = thin_border
    
    # الإجماليات
    total_row = 18 + len(data['items']) + 1
    
    sheet[f'F{total_row}'] = 'الإجمالي قبل الضريبة:'
    sheet[f'F{total_row}'].font = Font(name='Arial', size=11, bold=True)
    sheet[f'F{total_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet[f'G{total_row}'] = f'=SUM(G18:G{17+len(data["items"])})'
    sheet[f'G{total_row}'].font = Font(name='Arial', size=11, bold=True)
    sheet[f'G{total_row}'].number_format = '#,##0.00" ج.م"'
    sheet[f'G{total_row}'].border = thin_border
    sheet[f'G{total_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet[f'G{total_row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    sheet[f'F{total_row+1}'] = 'ضريبة القيمة المضافة (14%):'
    sheet[f'F{total_row+1}'].font = Font(name='Arial', size=11, bold=True)
    sheet[f'F{total_row+1}'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet[f'G{total_row+1}'] = f'=G{total_row}*0.14'
    sheet[f'G{total_row+1}'].font = Font(name='Arial', size=11, bold=True)
    sheet[f'G{total_row+1}'].number_format = '#,##0.00" ج.م"'
    sheet[f'G{total_row+1}'].border = thin_border
    sheet[f'G{total_row+1}'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet[f'G{total_row+1}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    sheet[f'F{total_row+2}'] = 'الإجمالي النهائي:'
    sheet[f'F{total_row+2}'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    sheet[f'F{total_row+2}'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet[f'F{total_row+2}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    sheet[f'G{total_row+2}'] = f'=G{total_row}+G{total_row+1}'
    sheet[f'G{total_row+2}'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    sheet[f'G{total_row+2}'].number_format = '#,##0.00" ج.م"'
    sheet[f'G{total_row+2}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    sheet[f'G{total_row+2}'].border = thin_border
    sheet[f'G{total_row+2}'].alignment = Alignment(horizontal='right', readingOrder=2)
    
    # ملاحظات
    notes_row = total_row + 4
    sheet[f'A{notes_row}'] = 'ملاحظات:'
    sheet[f'A{notes_row}'].font = Font(name='Arial', size=11, bold=True)
    sheet[f'A{notes_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet.merge_cells(f'A{notes_row+1}:G{notes_row+3}')
    sheet[f'A{notes_row+1}'] = data['notes']
    sheet[f'A{notes_row+1}'].border = thin_border
    sheet[f'A{notes_row+1}'].alignment = Alignment(vertical='top', wrap_text=True, horizontal='right', readingOrder=2)
    sheet[f'A{notes_row+1}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # التوقيعات
    sig_row = notes_row + 5
    sheet[f'A{sig_row}'] = 'المدير المسؤول'
    sheet[f'A{sig_row}'].font = Font(name='Arial', size=11, bold=True)
    sheet[f'A{sig_row}'].alignment = Alignment(horizontal='center', readingOrder=2)
    sheet.merge_cells(f'A{sig_row}:C{sig_row}')
    
    sheet[f'E{sig_row}'] = 'المورد'
    sheet[f'E{sig_row}'].font = Font(name='Arial', size=11, bold=True)
    sheet[f'E{sig_row}'].alignment = Alignment(horizontal='center', readingOrder=2)
    sheet.merge_cells(f'E{sig_row}:G{sig_row}')
    
    for i in range(1, 4):
        sheet[f'A{sig_row+i}'] = f'{"الاسم" if i==1 else "التوقيع" if i==2 else "التاريخ"}: _______________'
        sheet[f'A{sig_row+i}'].alignment = Alignment(horizontal='center', readingOrder=2)
        sheet.merge_cells(f'A{sig_row+i}:C{sig_row+i}')
        
        sheet[f'E{sig_row+i}'] = f'{"الاسم" if i==1 else "التوقيع" if i==2 else "التاريخ"}: _______________'
        sheet[f'E{sig_row+i}'].alignment = Alignment(horizontal='center', readingOrder=2)
        sheet.merge_cells(f'E{sig_row+i}:G{sig_row+i}')
    
    # ختم
    sheet[f'A{sig_row+4}'] = '(ختم الشركة)'
    sheet[f'A{sig_row+4}'].font = Font(name='Arial', size=9, italic=True)
    sheet[f'A{sig_row+4}'].alignment = Alignment(horizontal='center', readingOrder=2)
    sheet.merge_cells(f'A{sig_row+4}:C{sig_row+4}')
    
    sheet[f'E{sig_row+4}'] = '(ختم المورد)'
    sheet[f'E{sig_row+4}'].font = Font(name='Arial', size=9, italic=True)
    sheet[f'E{sig_row+4}'].alignment = Alignment(horizontal='center', readingOrder=2)
    sheet.merge_cells(f'E{sig_row+4}:G{sig_row+4}')
    
    # ضبط عرض الأعمدة
    sheet.column_dimensions['A'].width = 8
    sheet.column_dimensions['B'].width = 14
    sheet.column_dimensions['C'].width = 18
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['E'].width = 12
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 15
    
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[17].height = 25
    
    return wb

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/generate', methods=['POST'])
def generate():
    data = request.json
    wb = create_formatted_excel(data)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'طلب_توريد_{data["poNumber"]}.xlsx'
    )

def open_browser():
    """فتح المتصفح تلقائياً بعد ثانية واحدة"""
    time.sleep(1.5)
    webbrowser.open('http://localhost:5000')

if __name__ == '__main__':
    print("=" * 60)
    print("🚀 Supply Order Creation System - Al-Amana Company")
    print("=" * 60)
    print("✅ The system is now running...")
    print("🌐 Your browser will open automatically...")
    print("📍 URL: http://localhost:5000")
    print("=" * 60)
    print("⚠️  To stop the system: Press Ctrl+C")
    print("=" * 60)

    
    # فتح المتصفح في thread منفصل
    threading.Thread(target=open_browser, daemon=True).start()
    
    # تشغيل السيرفر
    app.run(debug=False, host='0.0.0.0', port=5000, use_reloader=False)
