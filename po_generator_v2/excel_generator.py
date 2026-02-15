# -*- coding: utf-8 -*-
"""
مولد ملفات Excel الاحترافية
Professional Excel File Generator
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_professional_excel(order_data):
    """
    إنشاء ملف Excel احترافي لطلب التوريد
    Create professional Excel file for purchase order
    """
    wb = Workbook()
    sheet = wb.active
    sheet.title = "طلب توريد"
    sheet.sheet_view.rightToLeft = True
    
    # تعريف الحدود - Define borders
    thin_border = Border(
        left=Side(style='thin', color='808080'),
        right=Side(style='thin', color='808080'),
        top=Side(style='thin', color='808080'),
        bottom=Side(style='thin', color='808080')
    )
    
    medium_border = Border(
        left=Side(style='medium', color='2B3E50'),
        right=Side(style='medium', color='2B3E50'),
        top=Side(style='medium', color='2B3E50'),
        bottom=Side(style='medium', color='2B3E50')
    )
    
    # تعريف الألوان الاحترافية - Define professional colors
    HEADER_COLOR = '2B3E50'  # أزرق داكن
    SECTION_HEADER_COLOR = '5B7A99'  # أزرق متوسط
    LIGHT_GRAY = 'F2F2F2'  # رمادي فاتح جداً
    DATA_HIGHLIGHT = 'E8F0F8'  # أزرق فاتح جداً
    TOTAL_COLOR = 'D9E2F3'  # أزرق فاتح للإجماليات
    FINAL_TOTAL_COLOR = '2B3E50'  # أزرق داكن للإجمالي النهائي
    
    # تعريف الخطوط - Define fonts
    TITLE_FONT = Font(name='Sakkal Majalla', size=22, bold=True, color='FFFFFF')
    SUBTITLE_FONT = Font(name='Sakkal Majalla', size=16, bold=True, color='FFFFFF')
    SECTION_FONT = Font(name='Traditional Arabic', size=14, bold=True, color='2B3E50')
    LABEL_FONT = Font(name='Traditional Arabic', size=12, bold=True)
    DATA_FONT = Font(name='Traditional Arabic', size=11)
    HEADER_TABLE_FONT = Font(name='Traditional Arabic', size=12, bold=True, color='FFFFFF')
    
    # === رأس الشركة - Company Header ===
    sheet.merge_cells('A1:G1')
    sheet['A1'] = 'شركة الأمانة للتوريدات العامة'
    sheet['A1'].font = TITLE_FONT
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
    sheet['A1'].fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type='solid')
    sheet.row_dimensions[1].height = 35
    
    sheet.merge_cells('A2:G2')
    sheet['A2'] = 'طلب توريد - Purchase Order'
    sheet['A2'].font = SUBTITLE_FONT
    sheet['A2'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
    sheet['A2'].fill = PatternFill(start_color=SECTION_HEADER_COLOR, end_color=SECTION_HEADER_COLOR, fill_type='solid')
    sheet.row_dimensions[2].height = 28
    
    # === بيانات الطلب الأساسية - Basic Order Info ===
    current_row = 4
    
    # رقم الطلب والتاريخ في صف واحد
    sheet[f'A{current_row}'] = 'رقم الطلب:'
    sheet[f'A{current_row}'].font = LABEL_FONT
    sheet[f'A{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    
    # Set column A width for labels
    sheet.column_dimensions['A'].width = 25
    
    sheet[f'B{current_row}'] = order_data.get('po_number', '')
    sheet[f'B{current_row}'].font = Font(name='Traditional Arabic', size=12, bold=True, color='0066CC')
    sheet[f'B{current_row}'].alignment = Alignment(horizontal='center', readingOrder=2)
    sheet[f'B{current_row}'].fill = PatternFill(start_color=DATA_HIGHLIGHT, end_color=DATA_HIGHLIGHT, fill_type='solid')
    sheet[f'B{current_row}'].border = thin_border
    
    # Set column B width
    sheet.column_dimensions['B'].width = 20
    
    sheet[f'D{current_row}'] = 'التاريخ:'
    sheet[f'D{current_row}'].font = LABEL_FONT
    sheet[f'D{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    
    # Set column D width
    sheet.column_dimensions['D'].width = 20
    
    sheet[f'E{current_row}'] = order_data.get('po_date', '')
    sheet[f'E{current_row}'].font = DATA_FONT
    sheet[f'E{current_row}'].alignment = Alignment(horizontal='center', readingOrder=2)
    sheet[f'E{current_row}'].fill = PatternFill(start_color=DATA_HIGHLIGHT, end_color=DATA_HIGHLIGHT, fill_type='solid')
    sheet[f'E{current_row}'].border = thin_border
    
    # Set column E width
    sheet.column_dimensions['E'].width = 20
    
    current_row += 1
    
    # الرقم الضريبي والسجل التجاري
    sheet[f'A{current_row}'] = 'الرقم الضريبي:'
    sheet[f'A{current_row}'].font = LABEL_FONT
    sheet[f'A{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    
    sheet[f'B{current_row}'] = order_data.get('company_tax_id', '')
    sheet[f'B{current_row}'].font = DATA_FONT
    sheet[f'B{current_row}'].alignment = Alignment(horizontal='center', readingOrder=2)
    sheet[f'B{current_row}'].fill = PatternFill(start_color=DATA_HIGHLIGHT, end_color=DATA_HIGHLIGHT, fill_type='solid')
    sheet[f'B{current_row}'].border = thin_border
    
    sheet[f'D{current_row}'] = 'السجل التجاري:'
    sheet[f'D{current_row}'].font = LABEL_FONT
    sheet[f'D{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    
    sheet[f'E{current_row}'] = order_data.get('commercial_reg', '')
    sheet[f'E{current_row}'].font = DATA_FONT
    sheet[f'E{current_row}'].alignment = Alignment(horizontal='center', readingOrder=2)
    sheet[f'E{current_row}'].fill = PatternFill(start_color=DATA_HIGHLIGHT, end_color=DATA_HIGHLIGHT, fill_type='solid')
    sheet[f'E{current_row}'].border = thin_border
    
    current_row += 2
    
    # === بيانات المورد - Supplier Information ===
    sheet.merge_cells(f'A{current_row}:G{current_row}')
    sheet[f'A{current_row}'] = 'بيانات المورد - Supplier Information'
    sheet[f'A{current_row}'].font = SECTION_FONT
    sheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
    sheet[f'A{current_row}'].fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
    sheet[f'A{current_row}'].border = thin_border
    sheet.row_dimensions[current_row].height = 25
    current_row += 1
    
    supplier = order_data.get('supplier', {})
    
    # اسم المورد
    sheet[f'A{current_row}'] = 'اسم المورد:'
    sheet[f'A{current_row}'].font = LABEL_FONT
    sheet[f'A{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet.merge_cells(f'B{current_row}:C{current_row}')
    sheet[f'B{current_row}'] = supplier.get('name', '')
    sheet[f'B{current_row}'].font = DATA_FONT
    sheet[f'B{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet[f'B{current_row}'].border = thin_border
    
    # Set column C width
    sheet.column_dimensions['C'].width = 25
    
    sheet[f'D{current_row}'] = 'الرقم الضريبي:'
    sheet[f'D{current_row}'].font = LABEL_FONT
    sheet[f'D{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet[f'E{current_row}'] = supplier.get('tax_id', '')
    sheet[f'E{current_row}'].font = DATA_FONT
    sheet[f'E{current_row}'].alignment = Alignment(horizontal='center', readingOrder=2)
    sheet[f'E{current_row}'].border = thin_border
    
    current_row += 1
    
    # تليفون وإيميل
    sheet[f'A{current_row}'] = 'التليفون:'
    sheet[f'A{current_row}'].font = LABEL_FONT
    sheet[f'A{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet[f'B{current_row}'] = supplier.get('phone', '')
    sheet[f'B{current_row}'].font = DATA_FONT
    sheet[f'B{current_row}'].alignment = Alignment(horizontal='center', readingOrder=2)
    sheet[f'B{current_row}'].border = thin_border
    
    sheet[f'D{current_row}'] = 'البريد الإلكتروني:'
    sheet[f'D{current_row}'].font = LABEL_FONT
    sheet[f'D{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet.merge_cells(f'E{current_row}:F{current_row}')
    sheet[f'E{current_row}'] = supplier.get('email', '')
    sheet[f'E{current_row}'].font = DATA_FONT
    sheet[f'E{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet[f'E{current_row}'].border = thin_border
    
    # Set column F and G widths
    sheet.column_dimensions['F'].width = 25
    sheet.column_dimensions['G'].width = 20
    
    current_row += 1
    
    # العنوان
    sheet[f'A{current_row}'] = 'العنوان:'
    sheet[f'A{current_row}'].font = LABEL_FONT
    sheet[f'A{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet.merge_cells(f'B{current_row}:G{current_row}')
    sheet[f'B{current_row}'] = supplier.get('address', '')
    sheet[f'B{current_row}'].font = DATA_FONT
    sheet[f'B{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
    sheet[f'B{current_row}'].border = thin_border
    
    current_row += 2
    
    # === جدول الأصناف - Items Table ===
    sheet.merge_cells(f'A{current_row}:G{current_row}')
    sheet[f'A{current_row}'] = 'أصناف الطلب - Order Items'
    sheet[f'A{current_row}'].font = SECTION_FONT
    sheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
    sheet[f'A{current_row}'].fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
    sheet[f'A{current_row}'].border = thin_border
    sheet.row_dimensions[current_row].height = 25
    current_row += 1
    
    # رأس الجدول - Table Headers
    headers = ['م', 'كود الصنف', 'اسم الصنف', 'الوصف', 'الكمية', 'سعر الوحدة', 'الإجمالي']
    header_widths = [25, 25, 35, 45, 18, 25, 25]
    
    for col_idx, (header, width) in enumerate(zip(headers, header_widths), start=1):
        cell = sheet.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = HEADER_TABLE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
        cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type='solid')
        cell.border = medium_border
        # Only set width if it's larger than what might already be set
        current_width = sheet.column_dimensions[get_column_letter(col_idx)].width
        if current_width is None or width > current_width:
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
    
    sheet.row_dimensions[current_row].height = 22
    current_row += 1
    
    # بيانات الأصناف - Items Data
    items = order_data.get('items', [])
    subtotal = 0.0
    
    for idx, item in enumerate(items, start=1):
        row_color = LIGHT_GRAY if idx % 2 == 0 else 'FFFFFF'
        
        # م - Number
        cell = sheet.cell(row=current_row, column=1)
        cell.value = idx
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
        cell.border = thin_border
        
        # كود الصنف - Product Code
        cell = sheet.cell(row=current_row, column=2)
        cell.value = item.get('product_code', '')
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
        cell.border = thin_border
        
        # اسم الصنف - Product Name
        cell = sheet.cell(row=current_row, column=3)
        cell.value = item.get('product_name', '')
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal='right', vertical='center', readingOrder=2, wrap_text=True)
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
        cell.border = thin_border
        
        # الوصف - Description
        cell = sheet.cell(row=current_row, column=4)
        cell.value = item.get('description', '')
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal='right', vertical='center', readingOrder=2, wrap_text=True)
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
        cell.border = thin_border
        
        # الكمية - Quantity
        cell = sheet.cell(row=current_row, column=5)
        cell.value = item.get('quantity', 0)
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        
        # سعر الوحدة - Unit Price
        cell = sheet.cell(row=current_row, column=6)
        cell.value = item.get('unit_price', 0)
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        
        # الإجمالي - Total
        item_total = item.get('total_price', 0)
        subtotal += item_total
        cell = sheet.cell(row=current_row, column=7)
        cell.value = item_total
        cell.font = Font(name='Traditional Arabic', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        
        sheet.row_dimensions[current_row].height = 20
        current_row += 1
    
    # === الإجماليات - Totals ===
    current_row += 1
    
    # المجموع الفرعي - Subtotal
    sheet.merge_cells(f'A{current_row}:F{current_row}')
    sheet[f'A{current_row}'] = 'المجموع الفرعي (قبل الضريبة):'
    sheet[f'A{current_row}'].font = LABEL_FONT
    sheet[f'A{current_row}'].alignment = Alignment(horizontal='right', vertical='center', readingOrder=2)
    sheet[f'A{current_row}'].fill = PatternFill(start_color=TOTAL_COLOR, end_color=TOTAL_COLOR, fill_type='solid')
    sheet[f'A{current_row}'].border = thin_border
    
    sheet[f'G{current_row}'] = subtotal
    sheet[f'G{current_row}'].font = Font(name='Traditional Arabic', size=12, bold=True)
    sheet[f'G{current_row}'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
    sheet[f'G{current_row}'].fill = PatternFill(start_color=TOTAL_COLOR, end_color=TOTAL_COLOR, fill_type='solid')
    sheet[f'G{current_row}'].border = medium_border
    sheet[f'G{current_row}'].number_format = '#,##0.00'
    sheet.row_dimensions[current_row].height = 22
    current_row += 1
    
    # الضريبة - Tax
    tax_rate = order_data.get('tax_rate', 14)
    tax_amount = subtotal * (tax_rate / 100)
    sheet.merge_cells(f'A{current_row}:F{current_row}')
    sheet[f'A{current_row}'] = f'ضريبة القيمة المضافة ({tax_rate}%):'
    sheet[f'A{current_row}'].font = LABEL_FONT
    sheet[f'A{current_row}'].alignment = Alignment(horizontal='right', vertical='center', readingOrder=2)
    sheet[f'A{current_row}'].fill = PatternFill(start_color=TOTAL_COLOR, end_color=TOTAL_COLOR, fill_type='solid')
    sheet[f'A{current_row}'].border = thin_border
    
    sheet[f'G{current_row}'] = tax_amount
    sheet[f'G{current_row}'].font = Font(name='Traditional Arabic', size=12, bold=True)
    sheet[f'G{current_row}'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
    sheet[f'G{current_row}'].fill = PatternFill(start_color=TOTAL_COLOR, end_color=TOTAL_COLOR, fill_type='solid')
    sheet[f'G{current_row}'].border = medium_border
    sheet[f'G{current_row}'].number_format = '#,##0.00'
    sheet.row_dimensions[current_row].height = 22
    current_row += 1
    
    # الإجمالي النهائي - Final Total
    final_total = subtotal + tax_amount
    sheet.merge_cells(f'A{current_row}:F{current_row}')
    sheet[f'A{current_row}'] = 'الإجمالي النهائي (شامل الضريبة):'
    sheet[f'A{current_row}'].font = Font(name='Traditional Arabic', size=13, bold=True, color='FFFFFF')
    sheet[f'A{current_row}'].alignment = Alignment(horizontal='right', vertical='center', readingOrder=2)
    sheet[f'A{current_row}'].fill = PatternFill(start_color=FINAL_TOTAL_COLOR, end_color=FINAL_TOTAL_COLOR, fill_type='solid')
    sheet[f'A{current_row}'].border = medium_border
    
    sheet[f'G{current_row}'] = final_total
    sheet[f'G{current_row}'].font = Font(name='Traditional Arabic', size=14, bold=True, color='FFFFFF')
    sheet[f'G{current_row}'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
    sheet[f'G{current_row}'].fill = PatternFill(start_color=FINAL_TOTAL_COLOR, end_color=FINAL_TOTAL_COLOR, fill_type='solid')
    sheet[f'G{current_row}'].border = medium_border
    sheet[f'G{current_row}'].number_format = '#,##0.00 "ج.م"'
    sheet.row_dimensions[current_row].height = 25
    current_row += 2
    
    # === شروط التوريد - Delivery Terms ===
    if order_data.get('delivery_period') or order_data.get('delivery_location') or order_data.get('payment_terms'):
        sheet.merge_cells(f'A{current_row}:G{current_row}')
        sheet[f'A{current_row}'] = 'شروط التوريد - Delivery Terms'
        sheet[f'A{current_row}'].font = SECTION_FONT
        sheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
        sheet[f'A{current_row}'].fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
        sheet[f'A{current_row}'].border = thin_border
        sheet.row_dimensions[current_row].height = 25
        current_row += 1
        
        if order_data.get('delivery_period'):
            sheet[f'A{current_row}'] = 'مدة التوريد:'
            sheet[f'A{current_row}'].font = LABEL_FONT
            sheet[f'A{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
            sheet.merge_cells(f'B{current_row}:G{current_row}')
            sheet[f'B{current_row}'] = order_data.get('delivery_period', '')
            sheet[f'B{current_row}'].font = DATA_FONT
            sheet[f'B{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2, wrap_text=True)
            sheet[f'B{current_row}'].border = thin_border
            current_row += 1
        
        if order_data.get('delivery_location'):
            sheet[f'A{current_row}'] = 'مكان التسليم:'
            sheet[f'A{current_row}'].font = LABEL_FONT
            sheet[f'A{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
            sheet.merge_cells(f'B{current_row}:G{current_row}')
            sheet[f'B{current_row}'] = order_data.get('delivery_location', '')
            sheet[f'B{current_row}'].font = DATA_FONT
            sheet[f'B{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2, wrap_text=True)
            sheet[f'B{current_row}'].border = thin_border
            current_row += 1
        
        if order_data.get('payment_terms'):
            sheet[f'A{current_row}'] = 'شروط الدفع:'
            sheet[f'A{current_row}'].font = LABEL_FONT
            sheet[f'A{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2)
            sheet.merge_cells(f'B{current_row}:G{current_row}')
            sheet[f'B{current_row}'] = order_data.get('payment_terms', '')
            sheet[f'B{current_row}'].font = DATA_FONT
            sheet[f'B{current_row}'].alignment = Alignment(horizontal='right', readingOrder=2, wrap_text=True)
            sheet[f'B{current_row}'].border = thin_border
            current_row += 1
    
    # === ملاحظات - Notes ===
    if order_data.get('notes'):
        current_row += 1
        sheet.merge_cells(f'A{current_row}:G{current_row}')
        sheet[f'A{current_row}'] = 'ملاحظات - Notes'
        sheet[f'A{current_row}'].font = SECTION_FONT
        sheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
        sheet[f'A{current_row}'].fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
        sheet[f'A{current_row}'].border = thin_border
        sheet.row_dimensions[current_row].height = 25
        current_row += 1
        
        sheet.merge_cells(f'A{current_row}:G{current_row}')
        sheet[f'A{current_row}'] = order_data.get('notes', '')
        sheet[f'A{current_row}'].font = DATA_FONT
        sheet[f'A{current_row}'].alignment = Alignment(horizontal='right', vertical='top', readingOrder=2, wrap_text=True)
        sheet[f'A{current_row}'].border = thin_border
        sheet.row_dimensions[current_row].height = 40
    
    # === التوقيعات - Signatures ===
    current_row += 3
    sheet.merge_cells(f'A{current_row}:C{current_row}')
    sheet[f'A{current_row}'] = 'توقيع المورد\n___________________'
    sheet[f'A{current_row}'].font = LABEL_FONT
    sheet[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2, wrap_text=True)
    sheet.row_dimensions[current_row].height = 40
    
    sheet.merge_cells(f'E{current_row}:G{current_row}')
    sheet[f'E{current_row}'] = 'توقيع الشركة\n___________________'
    sheet[f'E{current_row}'].font = LABEL_FONT
    sheet[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2, wrap_text=True)
    
    return wb
